#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 19 08:44:53 2020

@author: emec
"""
import os
import argparse
from argparse import Namespace
import IPython
import sys
sys.path.append('/frustum_framework/detection_2d/train/deteva/')
import numpy as np
from hist_lib import read_predictions,read_ground_truth,associate_mat,compare_stats
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.styles import Font

def read_log(log_dir):
    # Read the network parameters from log file
    f = open(log_dir,"r")
    lines = f.readlines() 
    params = lines[0]
    # Arguments that the network trained with
    log_args = eval(params)
    return log_args

def read_summary(sum_dir):
    f = open(sum_dir,"r")
    lines = f.readlines()
    res_dict = {}
    # read the summary file
    for line in lines:
        # Get the result type with the AP metrics
        key_raw,res_raw = line.split(':')
        # Split the raw result type
        key = key_raw.split(' ')[0]
        # Gen a new key wit hthe result type
        res_dict[key] = {}        
        # Read AP values
        _,easy,moderate,hard = res_raw.split(' ')
        # Record AP values acc to the diff level
        res_dict[key]['easy']=float(easy)
        res_dict[key]['moderate']=float(moderate)
        res_dict[key]['hard']=float(hard)
    return res_dict
        
def write_file(content,file_name):
    f = open(file_name,'w')
    f.write(content)
    f.close()       
def get_summary(log_dir,logs_dict):
    print('!!!!! Reading test in {}'.format(log_dir))
    log_name = os.path.basename(log_dir)
    logs_dict[log_name]={}
    # Get log file path
    log_path = os.path.join(log_dir,'log_train.txt')
    # Read the network parameters from log file
    log_args = read_log(log_path)
    # get the names of different drives 
    drive_names = get_folder_names(os.path.join(log_dir,'detection_results','data'))
    # Read the summary
    logs_dict[log_name]['drives']=[]
    for drive in drive_names:
        drive_id_str = os.path.basename(drive)
        try:
            float(drive_id_str) # To check if it is really name of a drive
            sum_path = os.path.join(log_dir,'detection_results','data',drive_id_str,'plot','summary.txt')
            sum_dict=read_summary(sum_path)
            logs_dict[log_name]['args']=log_args
            logs_dict[log_name][drive_id_str]={}
            logs_dict[log_name][drive_id_str]['results'] = sum_dict
            logs_dict[log_name]['drives'].append(drive_id_str)
        except Exception as exc:
            print('Failed in reading logs')
            print(exc)
            pass
    

def get_summary_drive(log_dir,logs_dict,drive_id):
    print('!!!!! Started test in {}'.format(log_dir))
    log_name = os.path.basename(log_dir)
    dir_name = os.path.basename(os.path.dirname(log_dir))
    logs_dict[log_name]={}
    logs_dict[log_name]['folder_name']=dir_name
    # Get log file path
    log_path = os.path.join(log_dir,'log_train.txt')
    # Read the network parameters from log file
    log_args = read_log(log_path)
    # get the names of different drives 
    drive_names = get_folder_names(os.path.join(log_dir,'detection_results','data'))
    # Read the summary
    logs_dict[log_name]['drives']=[]
    
    drive_id_str = '{:04d}'.format(drive_id)
    try:
        float(drive_id_str) # To check if it is really name of a drive
        sum_path = os.path.join(log_dir,'detection_results','data',drive_id_str,'plot','summary.txt')
        sum_dict=read_summary(sum_path)
        logs_dict[log_name]['args']=log_args
        logs_dict[log_name][drive_id_str]={}
        logs_dict[log_name][drive_id_str]['results'] = sum_dict
        logs_dict[log_name]['drives'].append(drive_id_str)
    except Exception as exc:
        print('Failed in reading logs')
        print(exc)
        pass


def get_folder_names(path):
    log_directories = [os.path.join(path,p) for p in os.listdir(path)\
                       if os.path.isdir(os.path.join(path,p))]
    return log_directories

def print_line(line_list):
    line_str = '|'
    for ent in line_list:
        line_str+='{:^25}|'.format(ent)
    return line_str

def print_row_sepa(line_list):
    line_str=''
    for ent in line_list:
        line_str+='|:'
        for i in range(23):
            line_str+='-'
        line_str+=':'
    line_str+='|'
    return line_str

def dump_dict(logs_dict,arg_list=['cos_loss','cos_loss_batch_thr','cos_loss_prop',\
                                  'cos_loss_weight','num_point','tau','track_features',\
                                  'track_net','tracks']):
   table = ''     
   header = print_line(['Log_name']+arg_list+['Drive ID']+['Easy','Moderate','Hard'])
   table+=header+'\n'
   separator= print_row_sepa(['Log_name']+arg_list+['Drive ID']+['Easy','Moderate','Hard'])
   table+=separator+'\n'
   
   log_names = list(logs_dict.keys())
   log_names.sort()
   for log_name in log_names:
       try:
           args_cls = logs_dict[log_name]['args']
       except:
           continue
       
       for drive_name in logs_dict[log_name]['drives']:
           logs_dict[log_name][drive_name]['cls'] = []
           print_list=[log_name]
           
           for arg in arg_list:
               print_list+=[eval('args_cls.{}'.format(arg))]
           
           print_list+= [drive_name]
           res = ['','',''] 
           ## Car
           try:
               car_res = logs_dict[log_name][drive_name]['results']['car_detection_3d']
               res[0]+='{:.1f}/'.format(car_res['easy'])
               res[1]+='{:.1f}/'.format(car_res['moderate'])
               res[2]+='{:.1f}/'.format(car_res['hard'])
               logs_dict[log_name][drive_name]['cls'].append('car')
           except:
               res[0]+='{:.1f}/'.format(0)
               res[1]+='{:.1f}/'.format(0)
               res[2]+='{:.1f}/'.format(0)
           ## pedestrian
           try:
               car_res = logs_dict[log_name][drive_name]['results']['pedestrian_detection_3d']
               res[0]+='{:.1f}/'.format(car_res['easy'])
               res[1]+='{:.1f}/'.format(car_res['moderate'])
               res[2]+='{:.1f}/'.format(car_res['hard'])
               logs_dict[log_name][drive_name]['cls'].append('pedestrian')
           except:
               res[0]+='{:.1f}/'.format(0)
               res[1]+='{:.1f}/'.format(0)
               res[2]+='{:.1f}/'.format(0)
           ## Cyclist
           try:
               car_res = logs_dict[log_name][drive_name]['results']['cyclist_detection_3d']
               res[0]+='{:.1f}/'.format(car_res['easy'])
               res[1]+='{:.1f}/'.format(car_res['moderate'])
               res[2]+='{:.1f}/'.format(car_res['hard'])
               logs_dict[log_name][drive_name]['cls'].append('cyclist')
           except:
               res[0]+='{:.1f}'.format(0)
               res[1]+='{:.1f}'.format(0)
               res[2]+='{:.1f}'.format(0)
           
           print_list+=res
           table+=print_line(print_list)
           table+='\n'
   return table

def calculate_histograms(log_dir,logs_dict,difficulty_levels,gt_root_dir):
    
    log_name = os.path.basename(log_dir)
    drive_ids = logs_dict[log_name]['drives']
    logs_dict[log_name]['full_dir'] = log_dir
    stat_results = {}
    for drive_id in drive_ids:
        print(log_name,drive_id)
        class_names = logs_dict[log_name][drive_id]['cls']
        drive_name = drive_id
        drive_id = int(drive_id)
        
        pred_folder = os.path.join(log_dir,'detection_results','data/{:04d}/data'.format(drive_id))
        gt_folder = os.path.join(gt_root_dir,'{:04d}.txt'.format(drive_id))
        pred = read_predictions(pred_folder,class_names)
        gts = read_ground_truth(gt_folder,class_of_interest=class_names,\
                                difficulty_include=difficulty_levels)
        stats= associate_mat(gts, pred,class_names,min_cost=0.3,max_cost=100)
        logs_dict[log_name][drive_name]['histograms'] ={}
        for cls in class_names:
            #stat_results[cls] = compare_stats(gts[-1], stats[-1], stats[-1], cls)[0]
            stat_results[cls] = []
            for fr_id in stats.best_prediction[cls].keys():
                stat_results[cls]+=stats.best_prediction[cls][fr_id]
            res1 = plt.hist(stat_results[cls], bins=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1])
            #IPython.embed()
            logs_dict[log_name][drive_name]['histograms'][cls]=res1[0]

def read_one_log(root_path,logs_dict):
    path = os.path.join(root_path,'log_train.txt') 
    f = open(path,"r")
    lines = f.readlines() 
    ## Find folder name
    params = lines[0]
    sp_params = params.split(',')
    for pr in sp_params:
        if pr[1:8]=='log_dir':
            folder_name = os.path.basename(pr[8:])[:-1]
    
    res_5 = [] # acc with iou 0.5
    res_7 = [] # acc with iou 0.7
    res = dict()        
    for i_line, line in enumerate(lines):
        if "EVALUATION" in line:
            try:
                res_7.append(float(lines[i_line+5].split(':')[-1]))
                res_5.append(float(lines[i_line+6].split(':')[-1]))
                eval_id = int(line.split(' ')[2])
                res[eval_id] = {'0.5':res_5[-1],'0.7':res_7[-1]}
            except Exception as e:
                print(e)
                print('Folder name {}'.format(root_path))
    
    i_max7 = np.argmax(res_7) ## ind of max value in iou 0.7
    max7 = np.max(res_7) # max accuracy in iou 0.7
    
    i_max5 = np.argmax(res_5) ## ind of max value in iou 0.5
    max5 = np.max(res_5) # max accuracy in iou 0.5
    f.close()
    try:
        logs_dict[folder_name]['accuracy'] = dict()
        logs_dict[folder_name]['accuracy']['accuracy_argmax_epoch_0.5']=i_max5
        logs_dict[folder_name]['accuracy']['accuracy_0.5']=max5
        logs_dict[folder_name]['accuracy']['accuracy_argmax_epoch_0.7']=i_max7
        logs_dict[folder_name]['accuracy']['accuracy_0.7']=max7
    except Exception as e:
        print(e)
    return [folder_name,i_max7,max7,i_max5,max5]


def float_with_comma(number,float_len=3):
    number = np.around(number,float_len)
    str_number = '{:}'.format(number)
    str_number = str_number.replace('.',',')
    return str_number
    

def write_to_excel(logs_dict, path_dir):
    '''
    histogram_dict : List of values of bins from histogram
    path : Path to log_dir
    
    '''
   	## Columns of the parameters
    cols = {'comp_name':1,
            'log_name':2,
            'param_id': 3,
            'folder_name':4,
            'tracks':5,
            'track_net':6,
            'track_features':7,
            'tau':8,
            'max_epoch':9,
            'num_point':10,
            'cos_loss':11,
            'cos_loss_weight':12,
            'cos_loss_batch_thr':13,
            'decay_step':14,
            'learning_rate':15,
            'batch_size':16,
            'two_branch':17,
            'only_whl':18,
            'temp_attention':19,
            'add_center':20,
            'output_attention':21,
            'time_indication':22,
            'accuracy_argmax_epoch_0.5':23,
            'accuracy_0.5':24,
            'accuracy_argmax_epoch_0.7':25,
            'accuracy_0.7':26
            }
    ## Calculate the maximum column used so far
    max_col = max([cols[k] for k in cols.keys()])
    ## place the APs to the columns 
    ## AP->  Drive1 (Easy, Mod, Hard), Drive2 (...) -> n_drive*3 columns
    cols['ap']=dict()
    
    dd = logs_dict['all_drives']
    dd.sort()
    logs_dict['all_drives'] = dd
    for d in logs_dict['all_drives']:
        cols['ap'][d]=dict()
        cols['ap'][d]['easy']=max_col+1
        cols['ap'][d]['moderate']=max_col+2
        cols['ap'][d]['hard']=max_col+3
        max_col = max_col+3
    
    ## Place histograms to the columns 
    ## histograms will only be written starting from 0.5-0.6 range
    ## |0.5-0.6(empty to separate)|Drive1|Drive2|Drive3|Drive4| -> n_hist*(drive+1) columns
    cols['hist']={0.5:{},0.6:{},0.7:{},0.8:{},0.9:{}}
    for k in cols['hist'].keys():
        cols['hist'][k]['empty']=max_col+1
        max_col+=1
        for d in logs_dict['all_drives']:
            cols['hist'][k][d]=max_col+1
            max_col+=1
    
    logs_dict['folder_name'] = os.path.basename(path_dir)
    ## Open the table sheet
    name_of_excel = os.path.join(path_dir,'all_summary.xlsx')
    #try:
    #    wb = openpyxl.load_workbook(name_of_excel)
    #    sheets={}
    #    for cl in logs_dict['all_cls']:
    #        sheets[cl] = wb.get_sheet_by_name(cl)
    #except:
    wb = openpyxl.Workbook()
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    sheets={}
    for cl in logs_dict['all_cls']:
        sheets[cl] = wb.create_sheet(title=cl)
    print(cols)
    ## Write first row (names of columns)
    for cl in logs_dict['all_cls']:
        for c_name in cols.keys():
            if c_name == 'ap':
                for d in logs_dict['all_drives']:
                    sheets[cl].cell(row=1,column=cols['ap'][d]['easy']).value = '{}-Easy'.format(d)
                    sheets[cl].cell(row=1,column=cols['ap'][d]['moderate']).value = '{}-Moderate'.format(d)
                    sheets[cl].cell(row=1,column=cols['ap'][d]['hard']).value = '{}-Hard'.format(d)
            elif c_name == 'hist':
                for k in cols['hist'].keys():
                    sheets[cl].cell(row=1,column=cols['hist'][k]['empty']).value = '{}-{} Gt-IoU'.format(k,k+0.1)
                    for d in logs_dict['all_drives']:
                        sheets[cl].cell(row=1,column=cols['hist'][k][d]).value = '{}'.format(d)
                        
            else:
                sheets[cl].cell(row=1,column=cols[c_name]).value = '{}'.format(c_name)
    
    for cl in logs_dict['all_cls']:
        for i_log,log_name in enumerate(logs_dict['log_dirs']):
            sheets[cl].cell(row=i_log+2,column=cols['log_name']).value = log_name
            for c_name in cols.keys():
                if c_name == 'ap':
                    
                    for d in logs_dict['all_drives']:
                        try:
                            name_field = '{}_detection_3d'.format(cl)
                            sheets[cl].cell(row=i_log+2,column=cols['ap'][d]['easy']).value = '{}'.format(float_with_comma(float(logs_dict[log_name][d]['results'][name_field]['easy'])))
                            sheets[cl].cell(row=i_log+2,column=cols['ap'][d]['moderate']).value = '{}'.format(float_with_comma(float(logs_dict[log_name][d]['results'][name_field]['moderate'])))
                            sheets[cl].cell(row=i_log+2,column=cols['ap'][d]['hard']).value = '{}'.format(float_with_comma(float(logs_dict[log_name][d]['results'][name_field]['hard'])))
                        except Exception as e:
                            print(e) 
                elif c_name == 'hist':
                    
                    for k in cols['hist'].keys():
                        for d in logs_dict['all_drives']:
                            try:
                                sheets[cl].cell(row=i_log+2,column=cols['hist'][k][d]).value = '{}'.format(int(logs_dict[log_name][d]['histograms'][cl][int(k*10)]))
                            except Exception as e:
                                print(e)
                elif 'accuracy' in c_name:
                    try:
                        sheets[cl].cell(row=i_log+2,column=cols[c_name]).value = '{}'.format(float_with_comma(logs_dict[log_name]['accuracy'][c_name]))
                    except Exception as e:
                        print(e)
                elif c_name == 'folder_name':
                    sheets[cl].cell(row=i_log+2,column=cols[c_name]).value = '{}'.format(logs_dict['folder_name'])
                else:
                    try:
                        sheets[cl].cell(row=i_log+2,column=cols[c_name]).value = '{}'.format(eval("logs_dict[log_name]['args'].{}".format(c_name)))
                    except:
                        pass
    
    print('Saving')                
    wb.save(name_of_excel)

       
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--log_dir',default=None,help='Path to a single log folder that contains log files of a training. This is overwritten by log_dirs if given.')
    parser.add_argument('--log_dirs',default=None,help='Path to log folders. Inside there should be log_* folders that contains logs of a training.')
    parser.add_argument('--gt_root_dir',default=None,help='Path to root dir of ground-truth tracking labels')
    parser.add_argument('--drive_id', type=int,default=-1,help='Drive id that the summary will be collected from')
    parser.add_argument('--difficulty', metavar='N', type=int, nargs='+',default=[0,1,2,3],help='By default ground-truth boxes from all difficulty levels are included. To see results on only some of the difficulty levels, give the difficulty id. 0: easy, 1:moderate, 2:hard, 3: unknown')
    
    args = parser.parse_args()
    
    ### All the information is written in this dictionary
    logs_dict = {}
    ### Read all the log directories in the base directory
    if args.log_dirs is not None:
        ## get names of directories in the base directory
        log_directories = get_folder_names(args.log_dirs)
    ### Read only one directory
    else:
        log_directories = [args.log_dir]
        
    for log_dir in log_directories:
        # If the evaluation is done. Otherwise it will fail
        try:
            # If there is no specific drive id given, this will look for evaluations of all drives
            if args.drive_id == -1:
                get_summary(log_dir,logs_dict)
            # If only one drive evaluation is asked
            else:
                get_summary_drive(log_dir,logs_dict,args.drive_id)
            
        except Exception as exc:
            del(logs_dict[os.path.basename(log_dir)])
            print('***** Failed in {}'.format(log_dir))
            print(exc)
    
    
    table=dump_dict(logs_dict)
    for log_dir in log_directories:
        if os.path.basename(log_dir) in logs_dict.keys():
            calculate_histograms(log_dir,logs_dict,args.difficulty,args.gt_root_dir)
    
    
    for log_dir in log_directories:
        read_one_log(log_dir,logs_dict)
    
    logs_dict['log_dirs'] = list(logs_dict.keys())
    logs_dict['all_drives'] = []
    logs_dict['all_cls'] = []
    for log_name in logs_dict['log_dirs']:
        logs_dict['all_drives']+=logs_dict[log_name]['drives']
        for dr in logs_dict[log_name]['drives']:
            logs_dict['all_cls']+=logs_dict[log_name][dr]['cls']
    logs_dict['all_drives'] = list(set(logs_dict['all_drives']))
    logs_dict['all_cls'] = list(set(logs_dict['all_cls']))
    write_to_excel(logs_dict,args.log_dirs)
    if args.log_dirs is None:
        args.log_dirs = args.log_dir
    write_to_excel(logs_dict,args.log_dirs)
    if args.drive_id == -1:
        write_file(table,os.path.join(args.log_dirs,'all_summary.txt'))
    else:
        write_file(table,os.path.join(args.log_dirs,'{:04d}_summary.txt'.format(args.drive_id)))
    print(table)
            
