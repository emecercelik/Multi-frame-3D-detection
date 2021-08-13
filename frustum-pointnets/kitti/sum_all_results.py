#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep  2 14:42:17 2020

@author: emec
"""
import argparse 

import os
import sys
sys.path.append('/frustum_framework/detection_2d/train/deteva/')
import numpy as np
from hist_lib import read_predictions,read_ground_truth,associate_mat,compare_stats
import matplotlib.pyplot as plt
import IPython

import openpyxl
from openpyxl.styles import Font

def write_to_excel(histogram_dict, path, eval_results):
    '''
    histogram_dict : List of values of bins from histogram
    path : Path to log_dir
    
    '''
    basename = os.path.basename(path)
    path_dir = os.path.dirname(path)
    name_of_excel = os.path.join(path_dir,'histogram_summary.xlsx')
    
    drives = list(histogram_dict.keys())
    classes = list(histogram_dict[drives[0]].keys())
    ## Column ids to keep AP values
    ap_cols=dict()
    ## Open if there is an excel sheet or create a new one
    try:
        wb = openpyxl.load_workbook(name_of_excel)
        sheets={}
        drive_cols = {}
        for cls in classes:
            ap_cols[cls] = dict()
            sheets[cls] = wb.get_sheet_by_name(cls)
            valid_col = 2
            for i in range(6):
                valid_col +=1
                drive_cols[0.4+i*0.1]={}
                for drive in drives:
                    drive_cols[0.4+i*0.1][drive] = valid_col
                    valid_col+=1
            for drive in drives:
                ap_cols[cls][drive] = dict()
                ap_cols[cls][drive]['easy'] = valid_col
                valid_col+=1
                ap_cols[cls][drive]['med'] = valid_col
                valid_col+=1
                ap_cols[cls][drive]['hard'] = valid_col
                valid_col+=1
    except:
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        sheets = {}
        drive_cols = {}
        for cls in classes:
            ap_cols[cls]=dict()
            sheet1 = wb.create_sheet(title=cls)
            sheet1.cell(row=1,column=1).value = 'Name'
            valid_col = 2
            for i in range(6):
                sheet1.cell(row=1,column=valid_col).value = 'Gt IoU {}-{}'.format(0.4+i*0.1,0.5+i*0.1)
                valid_col +=1
                drive_cols[0.4+i*0.1]={}
                for drive in drives:
                    sheet1.cell(row=1,column=valid_col).value = 'D{}'.format(drive)
                    drive_cols[0.4+i*0.1][drive] = valid_col
                    valid_col+=1
                    
            sheets[cls] = sheet1
            #IPython.embed()
            for drive in drives:
                ap_cols[cls][drive] = dict()
                sheet1.cell(row=1,column=valid_col).value = 'D{} AP Easy'.format(drive)
                ap_cols[cls][drive]['easy'] = valid_col
                valid_col+=1
                sheet1.cell(row=1,column=valid_col).value = 'D{} AP Med'.format(drive) 
                ap_cols[cls][drive]['med'] = valid_col
                valid_col+=1
                sheet1.cell(row=1,column=valid_col).value = 'D{} AP Hard'.format(drive) 
                ap_cols[cls][drive]['hard'] = valid_col
                valid_col+=1
                
    
    valid_rows = {}
    for cls in classes:
        valid_rows[cls] = sheets[cls].max_row+1
        for drive in drives:
            hist = histogram_dict[drive][cls]
            sheet = sheets[cls]
            sheet.cell(row=valid_rows[cls],column=1).value=basename
            print(basename)
            for i_h,h in enumerate(hist[4:]):
                sheet.cell(row=valid_rows[cls],column=drive_cols[0.4+i_h*0.1][drive]).value = h 
            #IPython.embed()
            sheet.cell(row=valid_rows[cls],column=ap_cols[cls][drive]['easy']).value = eval_results[drive][cls][0] 
            sheet.cell(row=valid_rows[cls],column=ap_cols[cls][drive]['med']).value = eval_results[drive][cls][1]
            sheet.cell(row=valid_rows[cls],column=ap_cols[cls][drive]['hard']).value = eval_results[drive][cls][2]
                
    IPython.embed()    
        
    wb.save(name_of_excel)

def read_evals(path):
    try:
        f = open(path,"r")
        lines = f.readlines()
        res=dict()
        for line in lines:
            if '3d' in line :
                res_list = line.split(' ')
                easy = float(res_list[2])
                med = float(res_list[3])
                hard = float(res_list[4])
                res[res_list[0].split('_')[0]] = [easy,med,hard]
        if 'car' not in res.keys():
            res['car'] = [-1,-1,-1]
        if 'pedestrian' not in res.keys():
            res['pedestrian'] = [-1,-1,-1]
        if 'cyclist' not in res.keys():
            res['cyclist'] = [-1,-1,-1]
        return res
    
    except Exception as E:
        print(E)
        res=dict()
        res['car'] = [-1,-1,-1]
        res['pedestrian'] = [-1,-1,-1]
        res['cyclist'] = [-1,-1,-1]
        return res
         
        
if __name__ == "__main__":
    
    parser = argparse.ArgumentParser()
    parser.add_argument('--class_names',metavar='N',nargs='+',help='Names of classes of interest. Ex: car pedestrian')
    parser.add_argument('--path_to_log',help='Path to the log folder of the frustum pointnet. The path should contain detection_results/data/<drive_id>/data folders.')
    parser.add_argument('--drive_ids', metavar='N', type=int, nargs='+',help='Drive IDs that will be evaluated to generate histogram.')
    parser.add_argument('--path_to_gts',help='Path that contains ground-truth track information in KITTI format. path/0011.txt etc.')
    parser.add_argument('--path_to_hist',default=None,help='Path where the histograms will be saved.')
    parser.add_argument('--inference', action='store_true', help='If given, the predictions will be checked under inference_results instead of detection_results.')
    parser.add_argument('--difficulty', metavar='N', type=int, nargs='+',default=[0,1,2,3],help='By default ground-truth boxes from all difficulty levels are included. To see results on only some of the difficulty levels, give the difficulty id. 0: easy, 1:moderate, 2:hard, 3: unknown')
    args = parser.parse_args()
    ## Get paths
    pred_folders=[]
    gt_folders=[]
    eval_paths=[]
    predictions=[]
    eval_results = {}
    gts=[] 
    stats = []
    stat_results = {}
    histogram_results = {}
    
    if args.inference:
        prediction_folder_name = 'inference_results'
    else:
        prediction_folder_name = 'detection_results'
    for drive_id in args.drive_ids:
        ### Folder that keep prediction txt files
        pred_folders.append(os.path.join(args.path_to_log,prediction_folder_name,'data/{:04d}/data'.format(drive_id)))
        ### Folder that keep gt txt files
        gt_folders.append(os.path.join(args.path_to_gts,'{:04d}.txt'.format(drive_id)))
        ### Paths to the evaluation summary
        eval_paths.append(os.path.join(args.path_to_log,prediction_folder_name,'data/{:04d}/plot/summary.txt'.format(drive_id)))
        ### Read predictions
        predictions.append(read_predictions(pred_folders[-1],args.class_names))
        ### Read ground-truths
        gts.append(read_ground_truth(gt_folders[-1],class_of_interest=args.class_names ,difficulty_include=args.difficulty))
        ### Associate gt with predictions to calculate statistics
        stats.append(associate_mat(gts[-1], predictions[-1],args.class_names,min_cost=0.3,max_cost=100))
        
        ###  Read eval results
        eval_results[drive_id] = read_evals(eval_paths[-1])
        
        ### Calculate histogram bins and plot
        histogram_results[drive_id]={}
        for cls in args.class_names:
            stat_results[cls] = compare_stats(gts[-1], stats[-1], stats[-1], cls)[0]
            plt.Figure()
            res1 = plt.hist(stat_results[cls], bins=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1])
            histogram_results[drive_id][cls]=res1[0]
            ax = plt.gca()
            for count, patch in zip(res1[0],res1[2]):
                ax.annotate(str(int(count)), xy=(patch.get_x()+0.025, patch.get_height()))
            #plt.ylim(-1,2500)
            plt.title('IoU values between predictions and ground-truth labels\nCls:{}, Drive:{}, Log:{}, Total # of Gt:{}'.format(cls,drive_id,os.path.basename(args.path_to_log),int(np.sum(res1[0]))))
            plt.xlabel('IoU with Gt')
            plt.ylabel('Num of predictions')
            fig_path = os.path.join(os.path.dirname(pred_folders[-1]),'hist.png')
            fig_dir = os.path.dirname(fig_path)
            if not os.path.isdir(fig_dir):
                os.makedirs(fig_dir)
            plt.savefig(fig_path)
            if args.path_to_hist is not None:
                plt.savefig(os.path.join(args.path_to_hist,'hist_{}_{}_{}.png'.format(os.path.basename(args.path_to_log),drive_id,cls)))
            plt.clf()
    write_to_excel(histogram_results,args.path_to_log,eval_results)  
            
            
        
        